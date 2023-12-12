from django.forms import modelform_factory
from django.http import HttpResponse
from django.shortcuts import render, redirect, get_object_or_404
from django.template import loader
from openpyxl.workbook import Workbook

from clientes.forms import ClienteFormulario
from clientes.models import Cliente


# Create your views here.
def agregar(request):
    pagina = loader.get_template('clientes/agregar.html')
    if request.method == 'GET':
        formulario = ClienteFormulario
    elif request.method == 'POST':
        formulario = ClienteFormulario(request.POST)
        if formulario.is_valid():
            formulario.save()
            return redirect('inicio')
    datos = {'formulario': formulario}
    return HttpResponse(pagina.render(datos, request))


def modificar(request, id):
    pagina = loader.get_template('clientes/modificar.html')
    cliente = get_object_or_404(Cliente, pk=id)
    if request.method == 'GET':
        formulario = ClienteFormulario(instance=cliente)
    elif request.method == 'POST':
        formulario = ClienteFormulario(request.POST, instance=cliente)
        if formulario.is_valid():
            formulario.save()
            return redirect('inicio')
    datos = {'formulario': formulario}
    return HttpResponse(pagina.render(datos, request))


def ver(request, id):
    cliente = get_object_or_404(Cliente, pk=id)
    datos = {'clientes': cliente}
    pagina = loader.get_template('clientes/ver.html')
    return render(request, 'clientes/ver.html', {'cliente': cliente})


def eliminar(request, id):
    cliente = get_object_or_404(Cliente, pk=id)
    if cliente:
        cliente.delete()
        return redirect('inicio')


# class ReportePersonasExcel():

    # Usamos el método get para generar el archivo excel
def generar_reporte(request, *args, **kwargs):
    # Obtenemos todas las personas de nuestra base de datos
    clientes = Cliente.objects.order_by('apellido', 'nombre')
    # Creamos el libro de trabajo
    wb = Workbook()
    # Definimos como nuestra hoja de trabajo, la hoja activa, por defecto la primera del libro
    ws = wb.active
    # En la celda B1 ponemos el texto 'REPORTE DE PERSONAS'
    ws['B1'] = 'REPORTE DE CLIENTES'
    # Juntamos las celdas desde la B1 hasta la E1, formando una sola celda
    ws.merge_cells('B1:G1')
    # Creamos los encabezados desde la celda B3 hasta la E3
    ws['B3'] = 'NOMBRE'
    ws['C3'] = 'APELLIDO'
    ws['D3'] = 'CORREO'
    ws['E3'] = 'TELEFONO'
    ws['F3'] = 'FECHA_PEDIDO'
    ws['G3'] = 'TIPO DE PRODUCTO'
    cont = 6
    # Recorremos el conjunto de personas y vamos escribiendo cada uno de los datos en las celdas
    for cliente in clientes:
        ws.cell(row=cont, column=2).value = cliente.nombre
        ws.cell(row=cont, column=3).value = cliente.apellido
        ws.cell(row=cont, column=4).value = cliente.correo
        ws.cell(row=cont, column=5).value = cliente.telefono
        ws.cell(row=cont, column=6).value = cliente.fecha_pedido
        #        ws.cell(row=cont, column=7).value = cliente.numero_de_pedido
        ws.cell(row=cont, column=7).value = cliente.tipo_producto.tipo_producto
        cont = cont + 1
        # Establecemos el nombre del archivo
    nombre_archivo = "ReporteClientesExcel.xlsx"
    # Definimos que el tipo de respuesta a devolver es un archivo de microsoft excel
    response = HttpResponse(content_type="application/ms-excel")
    contenido = "attachment; filename={0}".format(nombre_archivo)
    response["Content-Disposition"] = contenido
    wb.save(response)
    return response
